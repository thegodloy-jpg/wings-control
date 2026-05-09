"""生成模型×引擎支持矩阵 Excel 文件。"""
import os

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    import subprocess
    subprocess.check_call(["pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── 数据定义 ──
CATEGORIES = [
    {
        "name": "DeepSeek 推理/蒸馏系列",
        "models": [
            ["DeepSeek-R1-Distill-Qwen-1.5B", "Qwen2ForCausalLM", "✅", "✅", "✅", "✅", "Dense 小模型，四引擎无障碍"],
            ["DeepSeek-R1-Distill-Qwen-7B", "Qwen2ForCausalLM", "✅", "✅", "✅", "✅", "同上"],
            ["DeepSeek-R1-Distill-Qwen-14B", "Qwen2ForCausalLM", "✅", "✅", "✅", "✅", "同上"],
            ["DeepSeek-R1-Distill-Qwen-32B", "Qwen2ForCausalLM", "✅", "✅", "✅", "✅", "同上"],
            ["DeepSeek-R1-Distill-Llama-8B", "LlamaForCausalLM", "✅", "✅", "✅", "✅", "Llama 架构，全引擎原生支持"],
            ["DeepSeek-R1-Distill-Llama-70B", "LlamaForCausalLM", "✅", "✅", "✅", "✅", "同上，需多卡 TP"],
            ["DeepSeek-R1", "DeepseekV3ForCausalLM", "✅", "✅", "✅", "✅", "MoE 671B"],
            ["DeepSeek-R1-0528", "DeepseekV3ForCausalLM", "✅", "✅", "✅", "✅", "MindIE FC ✅ (deepseek_v3 parser)"],
        ]
    },
    {
        "name": "DeepSeek 通用/代码系列",
        "models": [
            ["DeepSeek-V3", "DeepseekV3ForCausalLM", "✅", "✅", "✅", "✅", "成熟支持"],
            ["DeepSeek-V3-0324", "DeepseekV3ForCausalLM", "✅", "✅", "✅", "✅", "MindIE FC ✅"],
            ["DeepSeek-V3.1", "DeepseekV3ForCausalLM", "✅", "✅", "✅🔧", "✅", "MindIE 必须配 deepseek_v31 parser"],
            ["DeepSeek-V3.2-EXP", "DeepseekV32ForCausalLM", "✅", "⚠️", "❌", "⚠️", "vLLM 0.19 已有 V3.2 支持；MindIE 无此架构"],
            ["DeepSeek-V4", "DeepseekV3ForCausalLM", "✅", "✅", "⚠️", "✅", "归为 V3 架构，MindIE 需实测"],
            ["DeepSeek-V4-280B", "DeepseekV3ForCausalLM", "✅", "⚠️", "⚠️", "⚠️", "280B 量级需多机 TP"],
            ["deepseek-coder236B", "DeepseekV3ForCausalLM", "✅", "✅", "⚠️", "✅", "Coder-V2 架构兼容 V3"],
        ]
    },
    {
        "name": "Qwen 通用文本系列",
        "models": [
            ["Qwen2.5-32B", "Qwen2ForCausalLM", "✅", "✅", "✅", "✅", "全引擎成熟支持"],
            ["QWQ-32B", "Qwen2ForCausalLM", "✅", "✅", "✅", "✅", "推理模型，同 Qwen2 架构"],
            ["Qwen3-8B/14B/32B", "Qwen3ForCausalLM", "✅", "✅", "✅", "✅", "MindIE qwen3 parser FC ✅"],
            ["Qwen3-30B-A3B", "Qwen3MoeForCausalLM", "✅", "✅", "✅", "⚠️", "MoE，MindIE FC ✅"],
            ["Qwen3-Next-80B-A3B", "Qwen3NextForCausalLM", "✅", "⚠️", "⚠️", "⚠️", "vLLM 0.19 有实现；其余待验证"],
            ["Qwen3-235B", "Qwen3MoeForCausalLM", "✅", "✅", "✅", "⚠️", "MoE 235B，MindIE FC ✅"],
            ["Qwen3.5-27B", "Qwen3_5ForConditionalGeneration", "✅", "⚠️", "❌", "⚠️", "vLLM 0.19 已支持；MindIE 无此架构"],
            ["Qwen3.5-35B-A3B", "Qwen3_5MoeForConditionalGeneration", "✅", "⚠️", "❌", "⚠️", "vLLM 0.19 有 config；MindIE ❌"],
            ["Qwen3.5-122B-A10B", "Qwen3_5MoeForConditionalGeneration", "✅", "⚠️", "❌", "⚠️", "同上"],
            ["Qwen3.5-397B-A17B", "Qwen3_5MoeForConditionalGeneration", "✅", "⚠️", "❌", "⚠️", "FC ❌"],
            ["Qwen3.5-397B-A17B-NVFP4", "Qwen3_5MoeForConditionalGeneration", "✅", "—", "❌", "✅", "NVFP4 仅 NVIDIA"],
        ]
    },
    {
        "name": "GLM 系列",
        "models": [
            ["GLM4-9B-0414", "GlmForCausalLM", "✅", "✅", "✅", "✅", "MindIE FC ✅ (chatglm4_9b)"],
            ["GLM4.7", "Glm4MoeForCausalLM", "✅", "⚠️", "⚠️", "⚠️", "vLLM 有 glm4_moe 实现；MindIE FC ❌"],
            ["GLM5", "GlmMoeDsaForCausalLM", "⚠️", "⚠️", "⚠️", "❌", "DSA 架构，vLLM 未明确列出；SGLang 无支持"],
            ["GLM-5.1", "GlmMoeDsaForCausalLM", "⚠️", "⚠️", "⚠️", "❌", "同 GLM5"],
        ]
    },
    {
        "name": "MiniMax 系列",
        "models": [
            ["MiniMax-M2.5", "MiniMaxM2ForCausalLM", "✅", "⚠️", "❌", "⚠️", "vLLM 官方列出 (TP ✅ PP ✅)"],
            ["MiniMax-M2.5-NVFP4", "MiniMaxM2ForCausalLM", "✅", "—", "❌", "⚠️", "NVFP4 仅 NVIDIA"],
            ["MiniMax-M2.5-SF-INT8", "MiniMaxM2ForCausalLM", "✅", "⚠️", "❌", "⚠️", "INT8 量化"],
            ["MiniMax-M2.7", "MiniMaxM2ForCausalLM", "✅", "⚠️", "❌", "⚠️", "同 M2.5 架构"],
        ]
    },
    {
        "name": "Kimi 系列",
        "models": [
            ["Kimi-K2-Thinking", "KimiK2ForCausalLM", "✅", "❌", "❌", "⚠️", "vLLM 有 reasoning/tool parser"],
        ]
    },
    {
        "name": "LLaMA 系列",
        "models": [
            ["LLaMA3-8B", "LlamaForCausalLM", "✅", "✅", "✅", "✅", "全引擎原生支持"],
        ]
    },
    {
        "name": "Embedding 模型",
        "models": [
            ["bge-large-zh-v1.5", "BertModel", "✅", "✅", "✅", "❌", "SGLang 不支持 Embedding"],
            ["bge-m3", "XLMRobertaModel", "✅", "✅", "✅", "❌", "同上"],
            ["Qwen3-Embedding-0.6B", "Qwen3ForCausalLM", "✅", "✅", "⚠️", "❌", "MindIE 池化策略需确认"],
        ]
    },
    {
        "name": "Reranker 模型",
        "models": [
            ["bge-reranker-large", "XLMRobertaForSeqCls", "✅", "✅", "✅", "❌", "SGLang 不支持 Rerank"],
            ["bge-reranker-v2-m3", "XLMRobertaForSeqCls", "✅", "✅", "✅", "❌", "同上"],
        ]
    },
    {
        "name": "视觉/多模态/全模态",
        "models": [
            ["Qwen2.5-VL-7B", "Qwen2_5VLForConditionalGeneration", "✅", "⚠️", "⚠️", "✅", "vLLM/SGLang 已支持 Qwen2.5-VL"],
            ["Qwen2.5-VL-72B", "Qwen2_5VLForConditionalGeneration", "✅", "⚠️", "⚠️", "✅", "同上，需多卡"],
            ["Qwen3-VL-8B", "Qwen3VLForConditionalGeneration", "✅", "⚠️", "❌", "⚠️", "vLLM 0.19 有 qwen3_vl 实现"],
            ["Qwen3-VL-30B", "Qwen3VLForConditionalGeneration", "✅", "⚠️", "❌", "⚠️", "同上"],
            ["Qwen3-VL-32B", "Qwen3VLForConditionalGeneration", "✅", "⚠️", "❌", "⚠️", "同上"],
            ["Qwen3vl-235b-Instruct", "Qwen3VLMoE", "✅", "❌", "❌", "⚠️", "超大 VL MoE"],
            ["Qwen3-Omni-30B-A3B", "Qwen3OmniMoeThinker", "✅", "❌", "❌", "⚠️", "vLLM 官方列出 (T+I+V+A)"],
            ["Qwen-Image", "待确认", "⚠️", "⚠️", "❌", "⚠️", "具体架构待确认"],
            ["混元video", "扩散模型", "—", "—", "—", "—", "非 LLM 推理引擎范畴"],
        ]
    },
]

HEADERS = ["模型名称", "架构", "vLLM 0.19", "vLLM-Ascend 0.18", "MindIE 2.30", "SGLang 5.9.0", "备注"]

# ── 颜色定义 ──
FILL_HEADER = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
FILL_CATEGORY = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
FILL_GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FILL_YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
FILL_RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FILL_GRAY = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
FILL_GREEN_WRENCH = PatternFill(start_color="A9D18E", end_color="A9D18E", fill_type="solid")

FONT_HEADER = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
FONT_CATEGORY = Font(name="微软雅黑", bold=True, size=11)
FONT_NORMAL = Font(name="微软雅黑", size=10)
FONT_STATUS = Font(name="微软雅黑", size=12)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def get_status_fill(val: str):
    """根据支持状态返回填充色。"""
    v = val.strip()
    if v.startswith("✅"):
        if "🔧" in v:
            return FILL_GREEN_WRENCH
        return FILL_GREEN
    if v.startswith("⚠️"):
        return FILL_YELLOW
    if v.startswith("❌"):
        return FILL_RED
    if v == "—":
        return FILL_GRAY
    return None


def build_workbook():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "模型×引擎支持矩阵"

    # 列宽
    col_widths = [32, 38, 14, 18, 14, 14, 50]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # 表头
    row = 1
    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER
    row += 1

    # 数据
    for cat in CATEGORIES:
        # 分类行
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(HEADERS))
        cell = ws.cell(row=row, column=1, value=cat["name"])
        cell.font = FONT_CATEGORY
        cell.fill = FILL_CATEGORY
        cell.alignment = ALIGN_LEFT
        for col in range(1, len(HEADERS) + 1):
            ws.cell(row=row, column=col).border = THIN_BORDER
            ws.cell(row=row, column=col).fill = FILL_CATEGORY
        row += 1

        for model_row in cat["models"]:
            for col, val in enumerate(model_row, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.border = THIN_BORDER
                if col in (3, 4, 5, 6):  # 引擎列
                    cell.font = FONT_STATUS
                    cell.alignment = ALIGN_CENTER
                    fill = get_status_fill(val)
                    if fill:
                        cell.fill = fill
                elif col == 1:
                    cell.font = Font(name="微软雅黑", bold=True, size=10)
                    cell.alignment = ALIGN_LEFT
                elif col == 2:
                    cell.font = Font(name="Consolas", size=9)
                    cell.alignment = ALIGN_LEFT
                else:
                    cell.font = FONT_NORMAL
                    cell.alignment = ALIGN_LEFT
            row += 1

    # ── 汇总 Sheet ──
    ws2 = wb.create_sheet("引擎能力概览")
    summary_headers = ["维度", "vLLM 0.19", "vLLM-Ascend 0.18", "MindIE 2.30", "SGLang 5.9.0"]
    summary_data = [
        ["文本 LLM 覆盖率", "~98%", "~70%", "~60%", "~65%"],
        ["MoE 模型", "全面支持", "部分 (DS/Qwen3 MoE ✅)", "DeepSeekV3 + Qwen3 MoE", "有限"],
        ["新架构 (Qwen3.5/V3.2)", "✅ 已适配", "⚠️ 滞后", "❌ 不支持", "⚠️ 需确认"],
        ["Embedding/Rerank", "✅", "✅", "✅", "❌"],
        ["VL 多模态", "广泛", "有限", "有限", "部分"],
        ["FC 工具调用", "最广 (含 V3.2)", "同 vLLM", "9 种 Processor", "有限"],
        ["核心短板", "无 Ascend 支持", "新架构滞后", "Qwen3.5/V3.2/MiniMax/Kimi 不支持", "无 Embedding，MoE 有限"],
    ]

    for col, h in enumerate(summary_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER

    for r, data in enumerate(summary_data, 2):
        for c, val in enumerate(data, 1):
            cell = ws2.cell(row=r, column=c, value=val)
            cell.font = FONT_NORMAL
            cell.alignment = ALIGN_CENTER if c > 1 else ALIGN_LEFT
            cell.border = THIN_BORDER

    ws2.column_dimensions["A"].width = 26
    for col_letter in ["B", "C", "D", "E"]:
        ws2.column_dimensions[col_letter].width = 30

    # ── 图例 Sheet ──
    ws3 = wb.create_sheet("图例")
    legend = [
        ["✅", "官方明确支持", FILL_GREEN],
        ["✅🔧", "支持但需特殊配置", FILL_GREEN_WRENCH],
        ["⚠️", "理论可用/需实测验证", FILL_YELLOW],
        ["❌", "不支持", FILL_RED],
        ["—", "类别不适用", FILL_GRAY],
    ]
    ws3.cell(row=1, column=1, value="符号").font = FONT_HEADER
    ws3.cell(row=1, column=1).fill = FILL_HEADER
    ws3.cell(row=1, column=1).border = THIN_BORDER
    ws3.cell(row=1, column=2, value="含义").font = FONT_HEADER
    ws3.cell(row=1, column=2).fill = FILL_HEADER
    ws3.cell(row=1, column=2).border = THIN_BORDER
    for r, (sym, desc, fill) in enumerate(legend, 2):
        c1 = ws3.cell(row=r, column=1, value=sym)
        c1.font = FONT_STATUS
        c1.fill = fill
        c1.alignment = ALIGN_CENTER
        c1.border = THIN_BORDER
        c2 = ws3.cell(row=r, column=2, value=desc)
        c2.font = FONT_NORMAL
        c2.alignment = ALIGN_LEFT
        c2.border = THIN_BORDER
    ws3.column_dimensions["A"].width = 12
    ws3.column_dimensions["B"].width = 30

    # 数据源说明
    ws3.cell(row=9, column=1, value="数据源").font = Font(name="微软雅黑", bold=True, size=11)
    sources = [
        "vLLM 0.19.0 官方文档: https://docs.vllm.ai/en/v0.19.0/models/supported_models.html",
        "MindIE 2.3.0 官方文档: https://www.hiascend.com/document/detail/zh/mindie/230/",
        "项目内部 model_utils.py 模型架构注册表",
        "项目内部 model-engine-function-call-analysis.md",
        f"生成日期: 2026-04-20",
    ]
    for i, src in enumerate(sources):
        ws3.cell(row=10 + i, column=1, value=src).font = FONT_NORMAL

    return wb


if __name__ == "__main__":
    output_path = os.path.join(os.path.dirname(__file__), "model_engine_support_matrix.xlsx")
    wb = build_workbook()
    wb.save(output_path)
    print(f"Excel 已生成: {output_path}")
